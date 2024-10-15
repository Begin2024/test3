# -*- coding: utf-8 -*-

import pandas as pd
import codecs
import numpy as np
import os
import pandas as pd
from datetime import datetime
import time
import openpyxl

# 记录程序开始时间
start_time = time.time()

addr_temp = '00000000'
base_addr = '0000_0000'
offset_addr = '0'

# 获取当前时间
current_time = datetime.now()

# 格式化输出当前时间
formatted_time = current_time.strftime("%Y-%m-%d %H:%M:%S")

Head_lines = [
'/********************************************************************',
f' # Author       : ---@insnex.com',
f' # Date         : {formatted_time} ',

f' # LastEditors  : ---insnex.com',
f' # LastEditTime : {formatted_time} ',
''' # FilePath     : reg_fpga.v
 # Description  : ---
 # Copyright (c) 2023 by Insnex.com, All Rights Reserved.
********************************************************************/
`include "define.h"
`include "revision.h"

module reg_fpga_temp(

    input wire                  clk                     ,
    input wire                  rst                     ,

    // Reg_fpga cmd
    input wire                  reg_req                 ,
    input wire                  reg_whrl                ,
    input wire [31:0]           reg_addr                ,
    input wire [31:0]           reg_wdata               ,
    output reg                  reg_ack                 ,
    output reg [31:0]           reg_rdata               ,

	output reg [0:0]            reg_soft_reset          ,
);




endmodule''']

Head = '\n'.join(Head_lines)

# 获取当前工作目录
current_directory = os.getcwd()
print("当前工作目录:", current_directory)

# 将工作目录移动到父目录
os.chdir("..")
os.chdir("..")

# 获取最新工作目录
now_directory = os.getcwd()
print("当前工作目录:", now_directory)

# 初始化变量以保存最新文件的信息
latest_file_path1 = None
latest_modification_time1 = 0
latest_file_path2 = None
latest_modification_time2 = 0


# 遍历当前工作目录及其所有子文件夹
for root, dirs, files in os.walk(now_directory):
    for filename in files:
        # 检查文件名是否包含'寄存器列表'且后缀为xlsx
        if '寄存器列表' in filename and filename.endswith('.xlsx'):
            # 获取文件的完整路径
            file_path1 = os.path.join(root, filename)
            print("找到匹配的文件:", filename)

            # 获取文件的修改日期
            modification_time1 = os.path.getmtime(file_path1)

            # 检查文件的修改日期是否比最新文件更新
            if modification_time1 > latest_modification_time1:
                latest_modification_time1 = modification_time1
                reg_list_path = file_path1


        # if 'reg_fpga_temp' in filename and filename.endswith('.v'):
        if filename == 'reg_fpga.v':
            # 获取文件的完整路径
            file_path2 = os.path.join(root, filename)
            print("找到匹配的文件:", filename)

            # 获取文件的修改日期
            modification_time2 = os.path.getmtime(file_path2)

            # 检查文件的修改日期是否比最新文件更新
            if modification_time2 > latest_modification_time2:
                latest_modification_time2 = modification_time2
                reg_fpga_path = file_path2

print("最新寄存器列表路径:", reg_list_path)
print("reg_fpga_temp路径:", reg_fpga_path)

# 输出最新文件的相对路径
if reg_fpga_path:
    relative_path = os.path.relpath(reg_fpga_path, now_directory)
    print("找到最新修改的匹配文件:", relative_path)

    # 新建或清空reg_fpga_temp.v文件
    temp_file_path = os.path.join(os.path.dirname(reg_fpga_path), 'reg_fpga_temp.v')
    with open(temp_file_path, 'w') as temp_file:
        # 写入你需要的内容
        temp_file.write(Head)
else:
    print("未找到匹配的文件")


def get_excel_max_row():
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(reg_list_path)

    # 选择第一个工作表
    sheet = workbook['寄存器列表']

    # 获取最大行数和列数
    max_row = sheet.max_row
    max_col = sheet.max_column

    # 逐行向上查找最后一行有数据的行
    last_row = max_row
    while last_row > 0:
        if any(sheet.cell(row=last_row, column=col).value for col in range(1, max_col + 1)):
            break
        last_row -= 1

    # 获取最后一行的数据
    last_row_data = [sheet.cell(row=max_row, column=col).value for col in range(1, max_col + 1)]

    # 打印最后一行的数据
    print(f"最后一行的数据：{last_row_data}")

    # 关闭 Excel 文件
    workbook.close()

    return last_row + 1

row_num = get_excel_max_row()

def get_excel_info(excel_row) :
    # 读取Excel文件
    data_frame  = pd.read_excel(reg_list_path, sheet_name='寄存器列表')

    # 获取整行的信息
    # excel_row   = int(excel_row)
    row_values  = data_frame.iloc[excel_row-2]  # 使得用户需要输入的行号和excel表中一致

    return row_values.tolist()[1:9]


# 代码插入函数
def reg_code_gen(write_info) :
    # 打开reg_fpga.v
    reg_fpga        = temp_file_path
    target_line     = 'endmodule'
    insert_content  = write_info

    with codecs.open(reg_fpga, "r") as file :
        lines       = file.readlines()

    target_line     = target_line.strip()

    # 找到 REG_READ这一行
    target_line_index = lines.index(target_line)

    # 在上一行插入
    lines.insert(target_line_index-1, insert_content)

    insert_content  = repr(insert_content)
    lines = [str(line) for line in lines]

    with open(reg_fpga, "w") as file :
        file.writelines(lines)

    return 0

def reg_code_port(port_info) :
    reg_fpga        = temp_file_path
    with open(reg_fpga, 'r') as file:
        content = file.read()

    # 找到第一个 ');' 前的最后一个非空格且非回车字符的位置
    index = content.find(');')
    if index != -1:
        index -= 1  # 减去一个字符以排除 ');' 前的空格

        while index >= 0 and content[index].isspace():  # 找到最后一个非空格字符
            index -= 1

    # 插入你想要的内容
    if index != -1:
        inserted_content = port_info
        content = content[:index + 1] + inserted_content + content[index + 1:]

    # 将修改后的内容写回文件
    with open(reg_fpga, 'w') as file:
        file.write(content)

# 删除最后一个端口的逗号
def reg_comma_delete() :
    reg_fpga        = temp_file_path
    # 读取文件内容
    with open(reg_fpga, 'r') as file:
        content = file.read()

    # 找到第一个 ');' 的位置
    index = content.find(');')

    # 如果找到了，则删除前面的最近一个 ','（包括可能的空格）
    if index != -1:
        last_comma_index = content[:index].rfind(',')
        if last_comma_index != -1:
            content = content[:last_comma_index].rstrip() + content[last_comma_index + 1:index] + content[index:]

    # 将修改后的内容写回文件
    with open(reg_fpga, 'w') as file:
        file.write(content)

#从excel数据中提取有用信息
def excel_cvt_reg_info(excel_row) :

    global base_addr
    global offset_addr
    global addr_temp

    # 获取数据
    cmd             = get_excel_info(excel_row)

    if (pd.isna(cmd[0]) == False) :
        base_addr       = cmd[0][0:9]
        # print(type(base_addr))
        # print('base_addr:', base_addr)
        addr_temp       = cmd[0][0:4] + cmd[0][5:9]
        # print(type(addr_temp))
        # print('addr_temp:', addr_temp)

    if (pd.isna(cmd[2]) == False) :
        offset_addr     = cmd[2]
        # print('offset_addr_type:',type(offset_addr))
        # print('offset_addr:', offset_addr)


    addr            = int(addr_temp, 16) + int(offset_addr, 16)
    # print(addr)
    addr_str        = hex(addr)[2:].upper()
    # print(addr_str)
    addr_z          = addr_str.zfill(8)
    # print(addr_z)
    addr_h        = "32'h" + addr_z[0:4] + "_" + addr_z[4:8]
    # print(addr_h)

    if (any(char.isupper() for char in str(cmd[3]))) :
        reg_name        = "`"+cmd[3]
    else :
        reg_name        = cmd[3]

    reg_type        = cmd[4]
    reg_field       = cmd[5]
    reg_width       = cmd[6]
    reg_value       = cmd[7]
    # 基地址， 偏移地址， 寄存器名称， 寄存器类型， 寄存器范围， 寄存器宽度， 寄存器默认值, 寄存器真地址
    reg_info        = base_addr, offset_addr, reg_name, reg_type, reg_field, reg_width, reg_value, addr_h

    return reg_info


# 寄存器信息转化为端口编码
def reg_info_cvt_port_code(write_info) :
    if write_info[3] == "RW" :
        code =         ("\n"
                        "    output reg [{}] {}{}{},"
        ).format(write_info[4], " " * (33 - len("    output reg [{}]".format(write_info[4])) - 2), write_info[2], " " * (57 - len(write_info[2]) - 33))
        return code

    if write_info[3] == "WC" :
        code =         ("\n"
                        "\toutput reg [{}] {}{}{},"
        ).format(write_info[4], " " * (33 - len("    output reg [{}]".format(write_info[4])) - 2), write_info[2], " " * (57 - len(write_info[2]) - 33))
        return code

    if write_info[3] == "RO" :
        code =         ("\n"
                        "\tinput wire [{}] {}{}{},"
        ).format(write_info[4], " " * (33 - len("    input wire [{}]".format(write_info[4])) - 2), write_info[2], " " * (57 - len(write_info[2]) - 33))
        return code

    if write_info[3] == "RC" :
        code =         ("\n"
                        "\tinput wire [{}] {}{}{},"
        ).format(write_info[4], " " * (33 - len("    input wire [{}]".format(write_info[4])) - 2), write_info[2], " " * (57 - len(write_info[2]) - 33))
        return code
    return print("type err")

# 寄存器信息转化为写逻辑编码
def reg_info_cvt_write_code(write_info) :
    if write_info[3] == "RW" :
        code = 		   ["    always @ (posedge clk or posedge rst) begin \n",
                        "        if (rst == 1\'b1) begin \n",
                        "            ",write_info[2]," <= ",write_info[5],write_info[6],";\n",
                        "        end else if (reg_req_r == 1\'b1 && reg_whrl_sync == 1\'b1 && reg_addr_sync == ",write_info[7],") begin\n",
                        "            ",write_info[2]," <= ","reg_wdata_sync[",write_info[4],"];","\n",
                        "        end else begin ","\n",
                        "            ",write_info[2]," <= ",write_info[2],";","\n",
                        "        end","\n",
                        "    end","\n\n"]
        return code

    if write_info[3] == "WC" :
        code = 		   ["    always @ (posedge clk or posedge rst) begin \n",
                        "        if (rst == 1\'b1) begin \n",
                        "            ",write_info[2]," <= ",write_info[5],write_info[6],";\n",
                        "        end else if (reg_req_r == 1\'b1 && reg_whrl_sync == 1\'b1 && reg_addr_sync == ",write_info[7],") begin\n",
                        "            ",write_info[2]," <= ","reg_wdata_sync[",write_info[4],"];","\n",
                        "        end else if (wc_clr_flag[0] == 1'b1) begin ","\n",
                        "            ",write_info[2]," <= ","\'d0"";","\n",
                        "        end else begin ","\n",
                        "            ",write_info[2]," <= ", write_info[2], ";\n"
                        "        end ", "\n"
                        "    end","\n\n"]
        return code
    if write_info[3] == "RO" :
        code = ["    //register ",write_info[2]," is Read only type","\n\n"]
        return code
    if write_info[3] == "RC" :
        code = ["    //register ",write_info[2]," is Read clr type","\n\n"]
        return code

    return print("type err")


# 寄存器信息转化为读逻辑编码
def reg_info_cvt_read_code(read_info) :
    if (read_info[3] == "RW" or read_info[3] == "RO" or read_info[3] == "WC") :
        # if (int(read_info[5]) < 32) :
        #     remain = int(32) - int(read_info[5])
        #     code    = ["                ",read_info[7],"\t : ","reg_rdata_pre[",read_info[4],"]\t <= {",remain,"\'h0, ",read_info[2],"};\n"]
        #     return code

        # code    = ["                ",read_info[7],"\t : ","reg_rdata_pre[",read_info[4],"]\t <= ",read_info[2],";\n"]
        code    = ["                ",read_info[7],"\t : ","reg_rdata_pre","\t <= ",read_info[2],";\n"]
        return code

    if (read_info[3] == "RC") :
        # 读清的FPGA逻辑是增加一个xx_temp， 如果reg_xx != 0  xx_temp <= reg_xx 如果读请求来了， xx_temp <= 'd0;
        # 后面有需要再写， 现在不着急弄
        code    = ["    //register ",read_info[2]," is Read clr type","\n\n"]

        return code

    return print("type err")

# 将一行信息转化为寄存器端口代码
def port_signal(excel_row_num) :
    reg_info = excel_cvt_reg_info(excel_row_num)
    judge = pd.isna(reg_info[4])
    if (judge == False) :
        reg_code = reg_info_cvt_port_code(reg_info)
        reg_code_str = ""
        for line in reg_code:
            if isinstance(line, str):
                   reg_code_str += line

        reg_code_port(reg_code_str)
    return 0

# 将一行信息转化为寄存器写操作代码代码
def write_signal(excel_row_num) :
    reg_info = excel_cvt_reg_info(excel_row_num)
    judge = pd.isna(reg_info[4])
    if (judge == False) :
        reg_code = reg_info_cvt_write_code(reg_info)
        reg_code_str = ""
        for line in reg_code:
            if isinstance(line, str):
                   reg_code_str += line

        reg_code_gen(reg_code_str)
    return 0

# 将一行信息转化为寄存器读操作代码代码
def read_signal(excel_row_num) :
    reg_info = excel_cvt_reg_info(excel_row_num)
    judge = pd.isna(reg_info[4])
    if (judge == False) :
        reg_code = reg_info_cvt_read_code(reg_info)
        reg_code_str = ""
        for line in reg_code:
            if isinstance(line, str):
                   reg_code_str += line

        reg_code_gen(reg_code_str)
    return 0

# 从excel第二行开始， 一直到最后一行

reg_start       = ["/********************************************************\n",
                   "*                        regs  Here                     *\n",
                   "********************************************************/\n",
                   "\n",
                   "    reg [2:0] \t\t\t\t\treg_req_dly \t\t\t;\n",
                   "    reg [1:0] \t\t\t\t\treg_whrl_dly \t\t\t;\n",
                   "    reg [63:0] \t\t\t\t\treg_addr_dly \t\t\t;\n",
                   "    reg [63:0] \t\t\t\t\treg_wdata_dly \t\t\t;\n",
                   "    reg [7:0] \t\t\t\t\treg_ack_high_cnt \t\t;\n",
                   "    reg [7:0] \t\t\t\t\treg_soft_reset_cnt \t\t;\n",
                   "    reg [31:0] \t\t\t\t\treg_rdata_pre \t\t\t;\n",
                   "    reg [31:0] \t\t\t\t\treg_test \t\t\t\t;\n",
                   "    reg [9:0] \t\t\t\t\twc_clr_flag \t\t\t;\n",
                   "\n",
                   "/********************************************************\n",
                   "*                        wires Here                     *\n",
                   "********************************************************/\n",
                   "\n",
                   "    wire \t\t\t\t\t\treg_req_r \t\t\t\t;\n",
                   "    wire \t\t\t\t\t\treg_whrl_sync \t\t\t;\n",
                   "    wire [31:0] \t\t\t\treg_addr_sync\t\t\t;\n",
                   "    wire [31:0] \t\t\t\treg_wdata_sync\t\t\t;\n",
                   "\n",
                   "/********************************************************\n",
                   "*                        logic Here                     *\n",
                   "********************************************************/\n",
                   "\n",
                   "    assign reg_req_r \t\t\t= reg_req_dly[2:1] == 2\'b01;\n",
                   "    assign reg_whrl_sync \t\t= reg_whrl_dly[1];\n",
                   "    assign reg_addr_sync \t\t= reg_addr_dly[63:32];\n",
                   "    assign reg_wdata_sync \t\t= reg_wdata_dly[63:32];\n",
                   "\n",
                   "    always @ (posedge clk) begin\n",
                   "        reg_req_dly <= {reg_req_dly[1:0], reg_req};\n",
                   "    end\n",
                   "\n",
                   "    always @ (posedge clk) begin\n",
                   "        reg_whrl_dly <= {reg_whrl_dly[0], reg_whrl};\n",
                   "    end\n",
                   "\n",
                   "    always @ (posedge clk) begin\n",
                   "        reg_addr_dly <= {reg_addr_dly[31:0], reg_addr};\n",
                   "    end\n",
                   "\n",
                   "    always @ (posedge clk) begin\n",
                   "        reg_wdata_dly <= {reg_wdata_dly[31:0], reg_wdata};\n",
                   "    end\n",
                   "\n",
                   "    always @ (posedge clk or posedge rst) begin\n",
                   "        if (rst == 1\'b1) begin\n",
                   "            reg_ack_high_cnt <= 8\'h0;\n",
                   "        end else if (reg_req_r == 1\'b1) begin\n",
                   "            reg_ack_high_cnt <= 8\'h1;\n",
                   "        end else if (reg_ack_high_cnt != 8\'h0) begin\n",
                   "            reg_ack_high_cnt <= reg_ack_high_cnt + 1\'b1;\n",
                   "        end else begin\n",
                   "            reg_ack_high_cnt <= reg_ack_high_cnt;\n",
                   "        end\n",
                   "    end\n",
                   "\n",
                   "    always @ (posedge clk or posedge rst) begin\n",
                   "        if (rst == 1\'b1) begin\n",
                   "            reg_ack <= 1\'h0;\n",
                   "        end else if (reg_ack_high_cnt == 8\'h2) begin\n",
                   "            reg_ack <= 1\'b1;\n",
                   "        end else begin\n",
                   "            reg_ack <= 1\'b0;\n",
                   "        end\n",
                   "    end\n",
                   "\n",
                   "    always @ (posedge clk or posedge rst) begin\n",
                   "        if (rst == 1'b1) begin\n",
                   "            reg_rdata <= 32\'d0;\n",
                   "        end else begin\n",
                   "            reg_rdata <= reg_rdata_pre;\n",
                   "        end\n",
                   "    end\n",
                   "\n",
                   "    always @ (posedge clk or posedge rst) begin\n",
                   "        if (rst == 1'b1) begin\n",
                   "            wc_clr_flag <= 10\'b00_0000_000;\n",
                   "        end else if (reg_req_r == 1\'b1) begin\n",
                   "            wc_clr_flag <= 10\'b10_0000_0000;\n",
                   "        end else begin\n",
                   "            wc_clr_flag <= wc_clr_flag >> 1\'b1;\n",
                   "        end\n",
                   "    end\n",
                   "\n"]


write_start     = "// Register write logic start--------------------------------------\n\n"
write_end       = "// Register write logic end----------------------------------------\n\n"
read_start      = "// Register read logic start---------------------------------------\n\n"
read_end        = "// Register read logic end-----------------------------------------\n\n"
read_head       =      ["\n    always @ (posedge clk or posedge rst) begin \n",
                        "        if (rst == 1\'b1) begin \n",
                        "            reg_rdata_pre <= 32\'d0;\n",
                        "        end else if (reg_req_r == 1\'b1) begin\n",
                        "            case (reg_addr_sync)\n"]

read_tail       =      ["                default          : reg_rdata_pre     <= 32\'h0;\n",
                        "            endcase\n",
                        "        end else begin\n",
                        "            reg_rdata_pre <= reg_rdata_pre;\n",
                        "        end\n",
                        "    end\n\n"]

# def recode(code) :
#     code_str = ""
#     for line in code :
#         if isinstance(line, str):
#             code_str != line

#     reg_code_gen(code_str)


def wr_logic():

    reg_start_str = ""
    for line in reg_start :
        if isinstance(line, str):
            reg_start_str += line

    reg_code_gen(reg_start_str)

    reg_code_gen(write_start)



    for i in range(11, row_num) :
        port_signal(i)

    reg_comma_delete()

    for i in range(2, row_num) :
        write_signal(i)

    reg_code_gen(write_end)
    reg_code_gen(read_start)

    read_head_str = ""
    for line in read_head :
        if isinstance(line, str):
            read_head_str += line

    reg_code_gen(read_head_str)

    for i in range(2, row_num) :
        read_signal(i)

    read_tail_str = ""
    for line in read_tail :
        if isinstance(line, str):
            read_tail_str += line

    reg_code_gen(read_tail_str)

    reg_code_gen(read_end)



wr_logic()

# 记录程序结束时间
end_time = time.time()

# 计算程序运行时间
elapsed_time = end_time - start_time

# 打印运行时间
print(f"程序运行时间：{elapsed_time} 秒")